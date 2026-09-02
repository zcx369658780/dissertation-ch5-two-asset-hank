"""Resumable year-subprocess scheduler for the Owner-run MP4C production batch."""
from __future__ import annotations
import argparse, csv, json, os, platform, subprocess, sys, time
from concurrent.futures import FIRST_COMPLETED, Future, ThreadPoolExecutor, wait
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any
import numpy as np
from openpyxl import Workbook

REPO_ROOT=Path(__file__).resolve().parents[2]
for root in (REPO_ROOT/"src",REPO_ROOT):
    if str(root) not in sys.path: sys.path.insert(0,str(root))
from validators.multi_province import mp4c_python_annual_empirical as empirical
from validators.multi_province import mp4c_python_annual_production as production
from validators.multi_province import mp4c_matlab_runtime_cache as runtime_cache
from validators.multi_province import mp4c_owner_a_2009_2022 as owner_a

YEARS=tuple(range(2009,2024)); DEFAULT_WORKERS=4

def sha(path:Path)->str:
    h=sha256();
    with path.open("rb") as f:
        for b in iter(lambda:f.read(1024*1024),b""): h.update(b)
    return h.hexdigest().upper()

def now()->str:return datetime.now(timezone.utc).isoformat()

def valid_success(year_dir:Path,year:int,canonical_sha:str,representation:str=runtime_cache.REPRESENTATION)->bool:
    marker=year_dir/"SUCCESS.json"
    if not marker.is_file(): return False
    try: p=json.loads(marker.read_text(encoding="utf-8"))
    except Exception:return False
    if p.get("schema")!="CH5_MP4C_YEAR_SUCCESS_V2" or p.get("representation")!=representation or p.get("year")!=year or p.get("status")!="SOURCE_CONVERGED" or p.get("runtime_input_sha256")!=canonical_sha or p.get("runtime_cache_sha256")!=runtime_cache.EXPECTED_CACHE_SHA256:return False
    if p.get("scientific_code_identities")!=production.scientific_identities() or p.get("checkpoint_schema")!=production.CHECKPOINT_SCHEMA:return False
    outputs=p.get("outputs",{})
    return bool(outputs) and all((year_dir/name).is_file() and sha(year_dir/name)==meta["sha256"] and (year_dir/name).stat().st_size==meta["bytes"] for name,meta in outputs.items())

def worker_env()->dict[str,str]:
    env=os.environ.copy(); env.update(production.THREAD_ENV); return env

def launch_year(year:int,canonical:Path,cache_path:Path,year_dir:Path)->dict[str,Any]:
    queued=now(); start=now(); tick=time.monotonic()
    proc=subprocess.run([sys.executable,str(REPO_ROOT/"validators/multi_province/mp4c_python_annual_production.py"),str(canonical),str(cache_path),str(year_dir)],cwd=REPO_ROOT,env=worker_env(),text=True)
    status="PASS" if proc.returncode==0 else ("FAIL" if proc.returncode==2 else "SHARED_FAIL")
    return {"year":year,"queued":queued,"start":start,"end":now(),"wall_clock_seconds":time.monotonic()-tick,"exit_code":proc.returncode,"status":status}

def prepare_runtime_inputs(data_root:Path,cache_path:Path,batch_root:Path,years:tuple[int,...]=YEARS,owner_a_inputs:bool=False)->dict[int,Path]:
    result={}
    for year in years:
        target=batch_root/"runtime_inputs"/str(year);path=target/f"calendar_{year}_matlab_runtime_cache_input.json"
        expected=owner_a.build_input(data_root,cache_path,year) if owner_a_inputs else runtime_cache.add_runtime_support(runtime_cache.load_runtime_year(cache_path,year),distance_workbook=empirical.primary_sources(data_root).distance_workbook,distance_sha256=empirical.SOURCE_HASHES["中国各省省会地理距离矩阵.xlsx"],max_sigmau=empirical.accepted_source_scalars().max_sigmau)
        expected_bytes=owner_a.canonical_bytes(expected) if owner_a_inputs else runtime_cache.canonical_bytes(expected)
        if target.exists():
            if not path.is_file() or path.read_bytes()!=expected_bytes:raise RuntimeError(f"shared runtime input defect for {year}")
        else:
            target.mkdir(parents=True,exist_ok=False);path.write_bytes(expected_bytes)
        result[year]=path
    return result

def aggregate(root:Path,statuses:list[dict[str,Any]],years:tuple[int,...])->None:
    passed=[s["year"] for s in statuses if s["status"] in ("PASS","SKIP")]
    rows=[]
    legacy_rows={}
    for year in passed:
        payload=json.loads((root/f"year_{year}"/"final_steady_state.json").read_text(encoding="utf-8"))
        for row in payload["final_31x20"]: rows.append({"year":year,**row})
        legacy_rows[year]=payload["legacy_workbook_rows"]
    suffix=f"{min(years)}_{max(years)}"
    with (root/f"steady_state_panel_{suffix}.csv").open("x",encoding="utf-8-sig",newline="") as f:
        w=csv.DictWriter(f,fieldnames=["year","name",*production.FINAL_FIELDS]);w.writeheader();w.writerows(rows)
    wb=Workbook();wb.remove(wb.active)
    legacy=("Yt0","Yt","Kt0","Kt","Lt0","Lt")
    # Source-backed analogue: fields unavailable in the frozen 31x20 use Yt/Kt/Lt only where named exactly.
    for field in legacy:
        ws=wb.create_sheet(f"稳态值_{field}");ws.append(["省份",*years])
        names=[r["name"] for r in rows if r["year"]==passed[0]] if passed else []
        for name in names:
            vals=[]
            for year in years:
                match=next((r for r in legacy_rows.get(year,[]) if r["name"]==name),None)
                vals.append(None if match is None else match[field])
            ws.append([name,*vals])
    wb.save(root/f"{suffix}_稳态值.xlsx")
    wb=Workbook();wb.remove(wb.active)
    for year in passed:
        ws=wb.create_sheet(f"{year}年"); terminal=json.loads((root/f"year_{year}"/"final_steady_state.json").read_text(encoding="utf-8"));order=terminal["province_order"]
        ws.append(order); matrix=np.load(root/f"year_{year}"/"Lt_mat_destination_row_origin_column.npy")
        for row in matrix:ws.append(row.tolist())
    wb.save(root/f"{suffix}_稳态Ltmat.xlsx")

def run_batch(data_root:Path,cache_path:Path,root:Path,workers:int,years:tuple[int,...]=YEARS,owner_a_inputs:bool=False)->int:
    if workers<1:raise ValueError("workers must be positive")
    if not years or tuple(sorted(set(years)))!=years: raise ValueError("years must be ordered and unique")
    if owner_a_inputs and years!=owner_a.YEARS: raise ValueError("Owner-A batch year set must be exactly 2009..2022")
    root.mkdir(parents=True,exist_ok=True); preflight_start=now(); overall=time.monotonic()
    contract=REPO_ROOT/"validators/multi_province/matlab_persistence_contract.json"; contract_copy=root/contract.name
    if contract_copy.exists():
        if sha(contract_copy)!=sha(contract): raise RuntimeError("shared MATLAB persistence contract defect")
    else: contract_copy.write_bytes(contract.read_bytes())
    if runtime_cache.file_sha256(cache_path)!=runtime_cache.EXPECTED_CACHE_SHA256:raise RuntimeError("shared MATLAB runtime cache identity defect")
    representation=owner_a.REPRESENTATION if owner_a_inputs else runtime_cache.REPRESENTATION
    manifest={"schema":"CH5_MP4C_FULL_BATCH_V3","representation":representation,"years":list(years),"workers":workers,"data_root":str(data_root.resolve()),"runtime_cache_path":str(cache_path.resolve()),"runtime_cache_sha256":runtime_cache.EXPECTED_CACHE_SHA256,"output_root":str(root.resolve()),"python":sys.executable,"thread_environment":production.THREAD_ENV,"start":now()}
    manifest_path=root/"batch_manifest.json"
    if manifest_path.exists():
        existing=json.loads(manifest_path.read_text(encoding="utf-8"))
        if existing.get("schema")!="CH5_MP4C_FULL_BATCH_V3" or existing.get("years")!=list(years) or existing.get("representation")!=representation or existing.get("data_root")!=str(data_root.resolve()) or existing.get("runtime_cache_sha256")!=runtime_cache.EXPECTED_CACHE_SHA256: raise RuntimeError("shared batch manifest defect")
        manifest=existing
    else: production.write_json_no_overwrite(manifest_path,manifest)
    canon=prepare_runtime_inputs(data_root,cache_path,root,years,owner_a_inputs); preflight_end=now(); statuses=[]; pending=[]
    for y in years:
        yd=root/f"year_{y}"; csha=sha(canon[y])
        if yd.exists():
            if valid_success(yd,y,csha,representation): statuses.append({"year":y,"status":"SKIP","queued":None,"start":None,"end":now(),"wall_clock_seconds":0,"exit_code":0})
            else: raise RuntimeError(f"incompatible existing output for {y}")
        else:pending.append(y)
    first_worker=None;last_worker=None
    shared_failure=False
    with ThreadPoolExecutor(max_workers=workers) as pool:
        active:dict[Future,int]={}
        while pending or active:
            while pending and len(active)<workers and not shared_failure:
                y=pending.pop(0);first_worker=first_worker or now();active[pool.submit(launch_year,y,canon[y],cache_path,root/f"year_{y}")]=y
            done,_=wait(active,return_when=FIRST_COMPLETED)
            for fut in done:
                item=fut.result();statuses.append(item);del active[fut];last_worker=now()
                if item["status"]=="SHARED_FAIL": shared_failure=True
            if shared_failure and not active and pending:
                statuses.extend({"year":y,"status":"BLOCKED_SHARED_FAILURE","queued":None,"start":None,"end":now(),"wall_clock_seconds":0,"exit_code":None} for y in pending);pending.clear()
            passed=sum(s["status"] in ("PASS","SKIP") for s in statuses);failed=sum(s["status"]=="FAIL" for s in statuses);running=sorted(active.values())
            print(f"[{len(statuses):02d}/{len(years)}] running={running} PASS={passed} FAIL={failed} elapsed={time.monotonic()-overall:.0f}s",flush=True)
    aggregate(root,statuses,years)
    statuses.sort(key=lambda x:x["year"])
    production.write_json_no_overwrite(root/"batch_summary.json",{"schema":"CH5_MP4C_BATCH_SUMMARY_V1","years":statuses})
    with (root/"batch_summary.csv").open("x",encoding="utf-8-sig",newline="") as f:
        w=csv.DictWriter(f,fieldnames=["year","status","queued","start","end","wall_clock_seconds","exit_code"]);w.writeheader();w.writerows(statuses)
    scientific_wall=None if not first_worker or not last_worker else (datetime.fromisoformat(last_worker)-datetime.fromisoformat(first_worker)).total_seconds()
    timing={"batch_start":manifest["start"],"batch_end":now(),"preflight_start":preflight_start,"preflight_end":preflight_end,"first_scientific_worker_start":first_worker,"last_scientific_worker_end":last_worker,"scientific_wall_clock_seconds":scientific_wall,"total_launcher_wall_clock_seconds":time.monotonic()-overall,"workers":workers,"python":sys.version,"numpy":np.__version__,"scipy":__import__("scipy").__version__,"os":platform.platform(),"logical_cpu_count":os.cpu_count(),"thread_environment":production.THREAD_ENV}
    production.write_json_no_overwrite(root/"batch_timing.json",timing)
    suffix=f"{min(years)}_{max(years)}"
    required=[root/n for n in ("batch_manifest.json","batch_summary.json","batch_summary.csv","batch_timing.json",f"steady_state_panel_{suffix}.csv",f"{suffix}_稳态值.xlsx",f"{suffix}_稳态Ltmat.xlsx","matlab_persistence_contract.json")]
    production.write_json_no_overwrite(root/"artifact_hash_manifest.json",{p.name:{"sha256":sha(p),"bytes":p.stat().st_size} for p in required})
    return 0 if all(s["status"] in ("PASS","SKIP") for s in statuses) else 2

def main(argv=None)->int:
    p=argparse.ArgumentParser();p.add_argument("--data-root",required=True);p.add_argument("--runtime-cache",required=True);p.add_argument("--output-root",required=True);p.add_argument("--workers",type=int,default=DEFAULT_WORKERS);p.add_argument("--years",type=int,nargs="+");p.add_argument("--owner-a-inputs",action="store_true");a=p.parse_args(argv)
    return run_batch(Path(a.data_root),Path(a.runtime_cache),Path(a.output_root),a.workers,tuple(a.years) if a.years else YEARS,a.owner_a_inputs)
if __name__=="__main__":raise SystemExit(main())
